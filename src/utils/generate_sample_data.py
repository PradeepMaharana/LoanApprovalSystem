"""
Generate 1000 rows of sample loan applicant data for Excel export
"""

import pandas as pd
from datetime import datetime, timedelta
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set random seed for reproducibility
random.seed(42)
np.random.seed(42)

# Define constants
EMPLOYMENT_TYPES = ["Salaried", "Self-Employed", "Freelancer", "Business Owner"]
LOCATIONS = [
    "New York, NY", "Los Angeles, CA", "Chicago, IL", "Houston, TX", "Phoenix, AZ",
    "Philadelphia, PA", "San Antonio, TX", "San Diego, CA", "Dallas, TX", "San Jose, CA",
    "Austin, TX", "Jacksonville, FL", "Fort Worth, TX", "Columbus, OH", "Charlotte, NC",
    "San Francisco, CA", "Indianapolis, IN", "Seattle, WA", "Denver, CO", "Boston, MA",
    "Miami, FL", "Portland, OR", "Atlanta, GA", "Las Vegas, NV", "Washington, DC",
    "Nashville, TN", "Detroit, MI", "Minneapolis, MN", "Memphis, TN", "Baltimore, MD"
]

def generate_applicant_data(num_rows: int = 1000) -> pd.DataFrame:
    """Generate sample loan applicant data"""

    data = {
        'Applicant ID': [],
        'Age': [],
        'Income': [],
        'Employment Type': [],
        'Credit Score': [],
        'Loan Amount': [],
        'Tenure (Months)': [],
        'Existing Liabilities': [],
        'Location': [],
        'Application Timestamp': []
    }

    # Base timestamp
    start_date = datetime.now() - timedelta(days=180)

    for i in range(1, num_rows + 1):
        # Applicant ID
        app_id = f"APP-{datetime.now().strftime('%Y')}-{i:06d}"
        data['Applicant ID'].append(app_id)

        # Age (18-75, with more weight towards 25-55)
        age = max(18, min(75, int(np.random.normal(45, 12))))
        data['Age'].append(age)

        # Income (30k - 500k, log-normal distribution)
        income = int(np.random.lognormal(10.5, 0.8))
        income = max(30000, min(500000, income))
        data['Income'].append(income)

        # Employment Type (weighted towards Salaried)
        employment_weights = [0.50, 0.25, 0.15, 0.10]
        employment_type = np.random.choice(EMPLOYMENT_TYPES, p=employment_weights)
        data['Employment Type'].append(employment_type)

        # Credit Score (300-850, normal distribution around 680)
        credit_score = int(np.random.normal(680, 80))
        credit_score = max(300, min(850, credit_score))
        data['Credit Score'].append(credit_score)

        # Loan Amount (based on income, typically 2-5x annual income)
        loan_multiplier = np.random.uniform(1.5, 5.5)
        loan_amount = int(income * loan_multiplier / 1000) * 1000  # Round to nearest 1000
        loan_amount = max(50000, min(2000000, loan_amount))
        # Make it multiple of 5000
        loan_amount = (loan_amount // 5000) * 5000
        data['Loan Amount'].append(loan_amount)

        # Tenure (months): 12-360, multiple of 12 (common terms)
        tenure_options = list(range(12, 361, 12))
        tenure = random.choice(tenure_options)
        data['Tenure (Months)'].append(tenure)

        # Existing Liabilities (0 - 50% of income)
        has_liabilities = random.random() < 0.7  # 70% have liabilities
        if has_liabilities:
            liabilities_ratio = np.random.uniform(0.05, 0.50)
            liabilities = int(income * liabilities_ratio / 1000) * 1000
        else:
            liabilities = 0
        data['Existing Liabilities'].append(liabilities)

        # Location
        location = random.choice(LOCATIONS)
        data['Location'].append(location)

        # Application Timestamp (spread over 180 days)
        days_offset = random.randint(0, 180)
        timestamp = start_date + timedelta(days=days_offset)
        # Add random hour, minute, second
        timestamp = timestamp.replace(
            hour=random.randint(8, 20),
            minute=random.randint(0, 59),
            second=random.randint(0, 59)
        )
        data['Application Timestamp'].append(timestamp)

    return pd.DataFrame(data)


def create_excel_file(df: pd.DataFrame, filename: str = "Loan_Applicants_Sample_Data.xlsx"):
    """Create formatted Excel file with sample data"""

    # Create Excel writer
    output_path = f"/home/ubuntu/Desktop/LoanApprovalSystem/{filename}"

    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Applicants', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Applicants']

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set column widths and format data
        column_widths = {
            'A': 15,  # Applicant ID
            'B': 8,   # Age
            'C': 12,  # Income
            'D': 16,  # Employment Type
            'E': 13,  # Credit Score
            'F': 13,  # Loan Amount
            'G': 16,  # Tenure (Months)
            'H': 18,  # Existing Liabilities
            'I': 16,  # Location
            'J': 22,  # Application Timestamp
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Format data rows
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), 2):
            for col_num, cell in enumerate(row, 1):
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Format specific columns
                if col_num == 2:  # Age
                    cell.alignment = Alignment(horizontal="center")
                elif col_num == 3:  # Income
                    cell.number_format = '$#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num == 5:  # Credit Score
                    cell.alignment = Alignment(horizontal="center")
                elif col_num == 6:  # Loan Amount
                    cell.number_format = '$#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num == 7:  # Tenure
                    cell.alignment = Alignment(horizontal="center")
                elif col_num == 8:  # Existing Liabilities
                    cell.number_format = '$#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num == 10:  # Timestamp
                    cell.number_format = 'mm/dd/yyyy hh:mm:ss'

        # Freeze top row
        worksheet.freeze_panes = "A2"

        # Add data validation info
        info_sheet = workbook.create_sheet('Info', 0)
        info_sheet.column_dimensions['A'].width = 50
        info_sheet.column_dimensions['B'].width = 30

        info_data = [
            ["Dataset Information", ""],
            ["", ""],
            ["Total Records", len(df)],
            ["Age Range", f"{df['Age'].min()} - {df['Age'].max()} years"],
            ["Income Range", f"${df['Income'].min():,} - ${df['Income'].max():,}"],
            ["Credit Score Range", f"{df['Credit Score'].min()} - {df['Credit Score'].max()}"],
            ["Loan Amount Range", f"${df['Loan Amount'].min():,} - ${df['Loan Amount'].max():,}"],
            ["Tenure Range", f"{df['Tenure (Months)'].min()} - {df['Tenure (Months)'].max()} months"],
            ["Liabilities Range", f"${df['Existing Liabilities'].min():,} - ${df['Existing Liabilities'].max():,}"],
            ["Unique Locations", len(df['Location'].unique())],
            ["", ""],
            ["Employment Type Distribution", "Count"],
            *[[emp_type, int(df[df['Employment Type'] == emp_type].shape[0])]
              for emp_type in sorted(df['Employment Type'].unique())],
        ]

        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = info_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12)
                elif row_idx == 12:
                    cell.font = Font(bold=True)
                    info_sheet.cell(row=row_idx, column=2).font = Font(bold=True)

    print(f"✅ Excel file created: {output_path}")
    return output_path


def print_summary(df: pd.DataFrame):
    """Print data summary statistics"""

    print("\n" + "="*80)
    print("SAMPLE DATA GENERATION SUMMARY")
    print("="*80)

    print(f"\n📊 Dataset Statistics:")
    print(f"  Total Records: {len(df):,}")
    print(f"  Date Range: {df['Application Timestamp'].min()} to {df['Application Timestamp'].max()}")

    print(f"\n👥 Age Statistics:")
    print(f"  Min: {df['Age'].min()} | Max: {df['Age'].max()} | Mean: {df['Age'].mean():.1f}")

    print(f"\n💰 Income Statistics:")
    print(f"  Min: ${df['Income'].min():,} | Max: ${df['Income'].max():,}")
    print(f"  Mean: ${df['Income'].mean():,.0f} | Median: ${df['Income'].median():,.0f}")

    print(f"\n📈 Credit Score Statistics:")
    print(f"  Min: {df['Credit Score'].min()} | Max: {df['Credit Score'].max()}")
    print(f"  Mean: {df['Credit Score'].mean():.1f} | Median: {df['Credit Score'].median():.1f}")

    print(f"\n🏦 Loan Amount Statistics:")
    print(f"  Min: ${df['Loan Amount'].min():,} | Max: ${df['Loan Amount'].max():,}")
    print(f"  Mean: ${df['Loan Amount'].mean():,.0f} | Median: ${df['Loan Amount'].median():,.0f}")

    print(f"\n📅 Tenure Statistics:")
    print(f"  Min: {df['Tenure (Months)'].min()} | Max: {df['Tenure (Months)'].max()}")
    print(f"  Mean: {df['Tenure (Months)'].mean():.1f} months")

    print(f"\n💳 Existing Liabilities Statistics:")
    print(f"  Min: ${df['Existing Liabilities'].min():,} | Max: ${df['Existing Liabilities'].max():,}")
    print(f"  Mean: ${df['Existing Liabilities'].mean():,.0f}")
    print(f"  Applicants with Liabilities: {(df['Existing Liabilities'] > 0).sum():,}")

    print(f"\n📍 Employment Type Distribution:")
    emp_dist = df['Employment Type'].value_counts()
    for emp_type, count in emp_dist.items():
        percentage = (count / len(df)) * 100
        print(f"  {emp_type}: {count:,} ({percentage:.1f}%)")

    print(f"\n🗺️ Location Distribution:")
    print(f"  Unique Locations: {df['Location'].nunique()}")
    print(f"  Top 5 Locations:")
    loc_dist = df['Location'].value_counts().head(5)
    for location, count in loc_dist.items():
        print(f"    {location}: {count}")

    print("\n" + "="*80 + "\n")


def main():
    """Main execution"""
    print("\n🚀 Starting Sample Data Generation...\n")

    # Generate data
    print("📝 Generating 1000 rows of sample data...")
    df = generate_applicant_data(1000)

    # Print summary
    print_summary(df)

    # Create Excel file
    print("💾 Creating Excel file with formatting...")
    excel_file = create_excel_file(df)

    print(f"✨ File saved at: {excel_file}")
    print(f"\n✅ Sample data generation complete!")
    print(f"\n📊 Data Preview (first 5 rows):")
    print(df.head().to_string(index=False))


if __name__ == "__main__":
    main()
